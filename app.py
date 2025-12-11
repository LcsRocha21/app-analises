import streamlit as st
import numpy as np
from PIL import Image
import pandas as pd
import sweetviz as sv
import seaborn as sns
import matplotlib.pyplot as plt
import plotly.express as px
import scipy.stats as stats
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
import prince  # Import ainda existe, mas não é mais usado na CA
from io import StringIO
from io import BytesIO
from adjustText import adjust_text
from matplotlib.lines import Line2D
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.tools.tools import add_constant
from matplotlib.colors import LinearSegmentedColormap
import io
from scipy.stats import pearsonr
import requests
from sklearn.preprocessing import LabelEncoder
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from matplotlib.colors import Normalize
from scipy.stats import chi2_contingency
import pyreadstat
import tempfile
import os
import matplotlib.pyplot as plt
from adjustText import adjust_text
import plotly.graph_objects as go



# Configurar página do Streamlit
st.set_page_config(page_title="Análise de Dados Automática", layout="wide")

# Adicionando um estilo CSS personalizado para alterar as cores
st.markdown(
    """
    <style>
    /* Fundo global com gradiente */
    .stApp {
        background: linear-gradient(to bottom, #ffffff, #81a53c);
        color: white;
    }

    /* Estilo para botões */
    button {
        background-color: #444;
        border: 1px solid white;
        border-radius: 5px;
        color: white;
        padding: 10px;
        cursor: pointer;
    }
    button:hover {
        background-color: #555;
    }

    /* Estilo para tabelas */
    .dataframe {
        background-color: #222;
        border: 1px solid #444;
        color: white;
    }

    /* Estilo para caixas de texto */
    textarea, input {
        background-color: #222;
        color: white;
        border: 1px solid #444;
    }
     /* Estilo para a tabela */
    .dataframe {
        background-color: #f5f5f5; /* Fundo da tabela (cinza claro) */
        border: 1px solid #ddd; /* Borda da tabela */
        color: #333; /* Texto da tabela */
        border-radius: 8px; /* Bordas arredondadas */
    }

    /* Estilo para as linhas alternadas */
    .dataframe tbody tr:nth-child(odd) {
        background-color: #e7f3e7; /* Verde claro para linhas ímpares */
    }
    .dataframe tbody tr:nth-child(even) {
        background-color: #ffffff; /* Branco para linhas pares */
    }

    /* Cabeçalho da tabela */
    .dataframe thead {
        background-color: #4caf50; /* Verde escuro */
        color: white; /* Texto branco */
    }
    </style>
    """,
    unsafe_allow_html=True
)

# URL da imagem
url = "https://institutoinforma.com.br/wp-content/uploads/2025/01/logo_informa.webp"

# Definir cabeçalhos personalizados
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36"
}

# Tentar baixar a imagem com o cabeçalho
response = requests.get(url, headers=headers)
response.raise_for_status()  # Gera um erro se a solicitação falhar

# Converter para um formato compatível com PIL
image_bytes = BytesIO(response.content)
logo = Image.open(image_bytes)

# Exibir no Streamlit
st.image(logo, width=300)


def carregar_dados(uploaded_file):
    """
    Carrega dados de CSV, Excel ou SPSS (SAV), preservando valores numéricos e associando labels apenas para exibição.
    """
    try:
        if uploaded_file.name.endswith('.csv'):
            try:
                dados = pd.read_csv(uploaded_file, encoding='utf-8')
            except UnicodeDecodeError:
                st.warning("Codificação 'utf-8' falhou. Tentando com 'latin1'.")
                dados = pd.read_csv(uploaded_file, encoding='latin1')

            st.success(f"Dados carregados com sucesso! Número de registros: {dados.shape[0]}")
            st.write("Visualização dos dados:", dados.head())
            # Para CSV, usamos o mesmo DF para dados e exibição
            return dados, dados

        elif uploaded_file.name.endswith('.xlsx'):
            dados = pd.read_excel(uploaded_file)
            st.success(f"Dados carregados com sucesso! Número de registros: {dados.shape[0]}")
            st.write("Visualização dos dados:", dados.head())
            return dados, dados  # Retorna o mesmo DataFrame para manter compatibilidade

        elif uploaded_file.name.endswith('.sav'):
            # Criar um arquivo temporário para leitura
            with tempfile.NamedTemporaryFile(delete=False, suffix=".sav") as tmp:
                tmp.write(uploaded_file.getvalue())
                tmp_path = tmp.name

            # Ler o arquivo SPSS preservando os valores numéricos
            dados, meta = pyreadstat.read_sav(tmp_path)

            # Remover o arquivo temporário após o carregamento
            os.unlink(tmp_path)

            # Criar um dicionário de mapeamento para labels
            label_dict = {}
            for var in meta.variable_value_labels:
                label_dict[var] = {int(k): v for k, v in meta.variable_value_labels[var].items()}  # Converter chaves para inteiros

            # Criar uma versão dos dados para exibição com labels
            dados_exibicao = dados.copy()
            for col in label_dict.keys():
                if col in dados.columns:  # Substituir apenas se a coluna existir
                    dados_exibicao[col] = dados[col].map(label_dict[col]).fillna(dados[col])  # Manter os valores numéricos caso não haja label

            st.success(f"Dados carregados com sucesso! Número de registros: {dados.shape[0]}")

            # Exibir os dados com labels sem afetar os cálculos
            st.write("Visualização com labels:", dados_exibicao.head())
            st.write("Dados numéricos usados nos cálculos:", dados.head())

            return dados, dados_exibicao  # Retorna ambas as versões

        else:
            st.error("Formato de arquivo não suportado. Use CSV, Excel ou SPSS (.sav).")
            return None, None

    except Exception as e:
        st.error(f"Erro ao carregar dados: {e}")
        return None, None


# Função para validar o arquivo carregado
def validar_arquivo(arquivo):
    """
    Função para validar se o arquivo carregado contém dados válidos.
    """
    if arquivo is not None:
        try:
            dados = pd.read_csv(arquivo)
            if dados.empty:
                st.error("O arquivo está vazio. Por favor, carregue um arquivo válido.")
                return None
            return dados
        except Exception as e:
            st.error(f"Erro ao ler o arquivo: {e}")
            return None
    else:
        st.warning("Nenhum arquivo carregado.")
        return None


def tratamento_dados(dados):
    st.header("Tratamento de Dados")

    if dados is None:
        st.warning("Nenhum dado disponível para tratamento. Por favor, carregue um arquivo primeiro.")
        return

    # Inicializar o estado da sessão para os dados tratados, se ainda não existir
    if 'dados_tratados' not in st.session_state:
        st.session_state['dados_tratados'] = dados.copy()

    # Usar os dados do estado da sessão
    dados_tratados = st.session_state['dados_tratados']

    # Exibir as primeiras linhas do DataFrame
    st.subheader("Visualização dos Dados")
    st.write(dados_tratados.head())

    # Seleção de coluna para tratamento
    coluna_para_tratar = st.selectbox("Escolha a coluna para modificar os valores", dados_tratados.columns)

    # Exibir valores únicos da coluna selecionada
    valores_unicos = dados_tratados[coluna_para_tratar].unique()
    st.write(f"Valores únicos na coluna '{coluna_para_tratar}':")
    st.write(valores_unicos)

    # Seleção do valor para modificar
    valor_para_alterar = st.selectbox(
        f"Selecione o valor para modificar na coluna '{coluna_para_tratar}'",
        valores_unicos
    )

    if st.button("Modificar valores para 9999"):
        if valor_para_alterar is not None:
            # Atualizar os dados substituindo o valor selecionado por 9999 na coluna escolhida
            st.session_state['dados_tratados'][coluna_para_tratar] = st.session_state['dados_tratados'][coluna_para_tratar].replace(valor_para_alterar, 9999)
            st.success(f"Todos os valores '{valor_para_alterar}' na coluna '{coluna_para_tratar}' foram substituídos por 9999.")
        else:
            st.warning("Nenhum valor selecionado para modificar.")

    # Exibir o DataFrame atualizado
    st.subheader("Dados Tratados")
    st.write(st.session_state['dados_tratados'])

    # Botão para salvar alterações em um arquivo
    if st.button("Salvar Alterações no Arquivo"):
        try:
            # Salvar os dados tratados como CSV com delimitador adequado e codificação ISO-8859-1
            csv = st.session_state['dados_tratados'].to_csv(index=False, sep=';', encoding='iso-8859-1')
            st.download_button(
                label="Baixar Dados Tratados",
                data=csv,
                file_name="dados_tratados.csv",
                mime="text/csv",
            )
            st.success("Alterações salvas e arquivo disponível para download!")
        except Exception as e:
            st.error(f"Erro ao salvar o arquivo: {e}")

    # Retornar os dados tratados
    return st.session_state['dados_tratados']


# Função para gráficos
def gerar_graficos(dados):
    st.header("Visualização de Dados")

    # Seletor de tipo de gráfico
    tipo_grafico = st.radio(
        "Selecione o tipo de gráfico:",
        ("Histograma", "Gráfico de Barras", "Gráfico de Dispersão", "Boxplot", "Gráfico de Pizza"),
    )

    # Seleção de colunas
    colunas = st.multiselect("Selecione as colunas para o gráfico:", dados.columns)

    # Gerar gráficos com base na seleção do usuário
    if len(colunas) > 0:
        if tipo_grafico == "Histograma":
            for coluna in colunas:
                st.subheader(f"Distribuição da coluna {coluna}")
                fig, ax = plt.subplots(figsize=(10, 6))
                sns.histplot(dados[coluna], kde=True, ax=ax)
                st.pyplot(fig)

        elif tipo_grafico == "Gráfico de Barras":
            for coluna in colunas:
                st.subheader(f"Gráfico de Barras para a coluna {coluna}")
                fig, ax = plt.subplots(figsize=(10, 6))
                dados[coluna].value_counts().plot(kind="bar", ax=ax, color="skyblue")
                ax.set_title(f"Distribuição de {coluna}")
                ax.set_xlabel("Categorias")
                ax.set_ylabel("Frequência")
                st.pyplot(fig)

        elif tipo_grafico == "Gráfico de Dispersão":
            if len(colunas) == 2:
                st.subheader(f"Gráfico de Dispersão: {colunas[0]} vs {colunas[1]}")

                fig, ax = plt.subplots(figsize=(10, 6))

                if not pd.api.types.is_numeric_dtype(dados[colunas[0]]) or not pd.api.types.is_numeric_dtype(dados[colunas[1]]):
                    st.info("Detectamos colunas categóricas. As categorias serão automaticamente mapeadas para valores numéricos.")

                sns.scatterplot(x=dados[colunas[0]], y=dados[colunas[1]], ax=ax)
                ax.set_title(f"{colunas[0]} vs {colunas[1]}")
                ax.set_xlabel(colunas[0])
                ax.set_ylabel(colunas[1])
                st.pyplot(fig)
            else:
                st.warning("Selecione exatamente 2 colunas para criar um gráfico de dispersão.")

        elif tipo_grafico == "Boxplot":
            for coluna in colunas:
                st.subheader(f"Boxplot para a coluna {coluna}")
                fig, ax = plt.subplots(figsize=(10, 6))
                sns.boxplot(y=dados[coluna], ax=ax, color="lightgreen")
                ax.set_title(f"Boxplot de {coluna}")
                st.pyplot(fig)

        elif tipo_grafico == "Gráfico de Pizza":
            for coluna in colunas:
                st.subheader(f"Gráfico de Pizza para a coluna {coluna}")
                dados_agrupados = dados[coluna].value_counts()
                fig = px.pie(
                    values=dados_agrupados.values,
                    names=dados_agrupados.index,
                    title=f"Distribuição de {coluna}",
                )
                st.plotly_chart(fig)
    else:
        st.warning("Selecione pelo menos uma coluna para gerar o gráfico.")


def preprocessar_dados(dados):
    for coluna in dados.columns:
        # Verificar se a coluna contém apenas números
        if dados[coluna].apply(lambda x: str(x).replace('.', '', 1).isdigit() if pd.notnull(x) else False).all():
            # Se todos os valores forem números, converter para numérico
            dados[coluna] = pd.to_numeric(dados[coluna], errors='coerce')
        else:
            # Se houver palavras, converter a coluna inteira para string
            dados[coluna] = dados[coluna].astype(str)

    return dados


# Função para gerar o relatório automatizado
def gerar_relatorio(dados):
    st.header("Relatório Automatizado com Sweetviz")

    # Pré-processando os dados para garantir que não haja tipos mistos
    dados = preprocessar_dados(dados)

    # Configuração das colunas a incluir no relatório
    st.subheader("Configurar Relatório")
    colunas_selecionadas = st.multiselect(
        "Selecione as colunas que deseja incluir no relatório:",
        options=dados.columns,
        default=dados.columns,
    )

    # Filtrar dados com base nas colunas selecionadas
    dados_filtrados = dados[colunas_selecionadas]

    # Tratar colunas de tipo misto
    for coluna in colunas_selecionadas:
        tipo_inferido = pd.api.types.infer_dtype(dados_filtrados[coluna])
        if tipo_inferido == 'mixed':
            st.warning(f"A coluna '{coluna}' possui tipos mistos. Convertendo para string.")
            dados_filtrados[coluna] = dados_filtrados[coluna].astype(str)

    if st.button("Gerar Relatório"):
        try:
            # Criar o relatório com os dados filtrados
            relatorio = sv.analyze(dados_filtrados)
            relatorio.show_html(filepath="relatorio_sweetviz.html")

            # Carregar o relatório gerado para exibir no Streamlit
            with open("relatorio_sweetviz.html", "r", encoding="utf-8") as f:
                html = f.read()
            st.components.v1.html(html, height=800, scrolling=True)

            st.success("Relatório gerado com sucesso!")
        except Exception as e:
            st.error(f"Erro ao gerar relatório: {e}")


# === CA MANUAL (SEM PRINCE) =====================================

def _ca_manual(tabela_contingencia, n_components=2):
    """
    Implementação manual de Análise de Correspondência via SVD,
    equivalente ao que o R faz.

    Retorna:
        coordenadas_linhas (DataFrame),
        coordenadas_colunas (DataFrame),
        eigenvalues (array)
    """
    N = tabela_contingencia.to_numpy(dtype=float)
    n = N.sum()
    if n <= 0:
        raise ValueError("Tabela de contingência vazia.")

    # Frequências relativas
    P = N / n
    r = P.sum(axis=1)   # massas das linhas
    c = P.sum(axis=0)   # massas das colunas

    # Remove linhas/colunas com massa zero
    row_mask = r > 0
    col_mask = c > 0

    if row_mask.sum() < 2 or col_mask.sum() < 2:
        raise ValueError("Linhas/colunas insuficientes com massa > 0 para a CA.")

    P_rc = P[np.ix_(row_mask, col_mask)]
    r_rc = r[row_mask]
    c_rc = c[col_mask]

    # Matrizes diagonais inversa da raiz
    D_r_inv_sqrt = np.diag(1.0 / np.sqrt(r_rc))
    D_c_inv_sqrt = np.diag(1.0 / np.sqrt(c_rc))

    expected = np.outer(r_rc, c_rc)
    S = D_r_inv_sqrt @ (P_rc - expected) @ D_c_inv_sqrt

    U, singvals, VT = np.linalg.svd(S, full_matrices=False)
    eigenvalues = singvals ** 2

    # Coordenadas principais
    F = D_r_inv_sqrt @ U @ np.diag(singvals)
    G = D_c_inv_sqrt @ VT.T @ np.diag(singvals)

    k = min(n_components, F.shape[1])
    F = F[:, :k]
    G = G[:, :k]
    eigenvalues = eigenvalues[:k]

    row_index = tabela_contingencia.index[row_mask]
    col_index = tabela_contingencia.columns[col_mask]
    cols = [f"Dim{i+1}" for i in range(k)]

    coordenadas_linhas = pd.DataFrame(F, index=row_index, columns=cols)
    coordenadas_colunas = pd.DataFrame(G, index=col_index, columns=cols)

    return coordenadas_linhas, coordenadas_colunas, eigenvalues

def analise_correspondencia(dados):
    st.header("Análise de Correspondência")

    # Seleção de colunas categóricas
    colunas_categoricas = st.multiselect(
        "Selecione as colunas categóricas para a análise:",
        dados.columns
    )

    if len(colunas_categoricas) < 2:
        st.warning("Selecione pelo menos duas colunas categóricas para realizar a análise.")
        return

    # Tabela de contingência
    tabela_contingencia = pd.crosstab(
        dados[colunas_categoricas[0]],
        dados[colunas_categoricas[1]],
        normalize=False
    )

    # CA manual
    try:
        coordenadas_linhas, coordenadas_colunas, eigenvalues = _ca_manual(
            tabela_contingencia,
            n_components=2
        )
    except Exception as e:
        st.error(f"Não foi possível calcular a Análise de Correspondência: {e}")
        return

    # Inverter sinais para ficar similar ao R
    coordenadas_linhas *= -1
    coordenadas_colunas *= -1

    if coordenadas_linhas.shape[1] < 2 or coordenadas_colunas.shape[1] < 2:
        st.error("A análise de correspondência não conseguiu gerar duas dimensões. Verifique os dados selecionados.")
        return

    # Inércia explicada
    st.subheader("Inércia explicada (variância)")
    explained_inertia = eigenvalues / eigenvalues.sum()
    for i, valor in enumerate(explained_inertia):
        st.write(f"Dim {i+1}: {valor * 100:.2f}%")

    # ---------- ESTADO (rótulos e deslocamentos) ----------

    if 'rotulos_linhas' not in st.session_state:
        st.session_state['rotulos_linhas'] = {i: str(i) for i in coordenadas_linhas.index}
    else:
        for i in coordenadas_linhas.index:
            st.session_state['rotulos_linhas'].setdefault(i, str(i))

    if 'rotulos_colunas' not in st.session_state:
        st.session_state['rotulos_colunas'] = {i: str(i) for i in coordenadas_colunas.index}
    else:
        for i in coordenadas_colunas.index:
            st.session_state['rotulos_colunas'].setdefault(i, str(i))

    if 'deslocamentos_linhas' not in st.session_state:
        st.session_state['deslocamentos_linhas'] = {i: (0.0, 0.0) for i in coordenadas_linhas.index}
    else:
        for i in coordenadas_linhas.index:
            st.session_state['deslocamentos_linhas'].setdefault(i, (0.0, 0.0))

    if 'deslocamentos_colunas' not in st.session_state:
        st.session_state['deslocamentos_colunas'] = {i: (0.0, 0.0) for i in coordenadas_colunas.index}
    else:
        for i in coordenadas_colunas.index:
            st.session_state['deslocamentos_colunas'].setdefault(i, (0.0, 0.0))

    # Legendas gerais
    legenda_linhas = st.text_input("Legenda para Linhas:", "Linhas")
    legenda_colunas = st.text_input("Legenda para Colunas:", "Colunas")

    # Controle para mostrar/ocultar itens de legenda
    mostrar_legenda = st.checkbox("Mostrar legenda (Linhas/Colunas)", value=False)

    # --------- EDIÇÃO NUMÉRICA DOS RÓTULOS / DESLOCAMENTOS ---------

    editar_rotulos = st.checkbox("Editar rótulos e deslocamentos")

    if editar_rotulos:
        st.subheader("Editar Rótulos e Posições")

        st.markdown("**Linhas**")
        for i in coordenadas_linhas.index:
            st.session_state['rotulos_linhas'][i] = st.text_input(
                f"Novo rótulo para linha '{i}':",
                value=st.session_state['rotulos_linhas'][i],
                key=f"rotulo_linha_{i}"
            )

            dx, dy = st.session_state['deslocamentos_linhas'][i]
            dx = st.number_input(
                f"Deslocamento X para linha '{i}':",
                min_value=-1.0,
                max_value=1.0,
                value=float(dx),
                step=0.01,
                format="%.2f",
                key=f"num_desloc_x_linha_{i}"
            )
            dy = st.number_input(
                f"Deslocamento Y para linha '{i}':",
                min_value=-1.0,
                max_value=1.0,
                value=float(dy),
                step=0.01,
                format="%.2f",
                key=f"num_desloc_y_linha_{i}"
            )
            st.session_state['deslocamentos_linhas'][i] = (dx, dy)

            if st.checkbox(f"Remover ponto e rótulo da linha '{i}'", key=f"remover_linha_{i}"):
                st.session_state['rotulos_linhas'][i] = None

        st.markdown("**Colunas**")
        for i in coordenadas_colunas.index:
            st.session_state['rotulos_colunas'][i] = st.text_input(
                f"Novo rótulo para coluna '{i}':",
                value=st.session_state['rotulos_colunas'][i],
                key=f"rotulo_coluna_{i}"
            )

            dx, dy = st.session_state['deslocamentos_colunas'][i]
            dx = st.number_input(
                f"Deslocamento X para coluna '{i}':",
                min_value=-1.0,
                max_value=1.0,
                value=float(dx),
                step=0.01,
                format="%.2f",
                key=f"num_desloc_x_coluna_{i}"
            )
            dy = st.number_input(
                f"Deslocamento Y para coluna '{i}':",
                min_value=-1.0,
                max_value=1.0,
                value=float(dy),
                step=0.01,
                format="%.2f",
                key=f"num_desloc_y_coluna_{i}"
            )
            st.session_state['deslocamentos_colunas'][i] = (dx, dy)

            if st.checkbox(f"Remover ponto e rótulo da coluna '{i}'", key=f"remover_coluna_{i}"):
                st.session_state['rotulos_colunas'][i] = None

    # ----------------- PRÉ-CÁLCULO DOS PONTOS VISÍVEIS -----------------

    max_deslocamento = 0.05
    pontos_linhas = []
    pontos_colunas = []

    for i, row in coordenadas_linhas.iterrows():
        if st.session_state['rotulos_linhas'][i] is None:
            continue
        x = row.iloc[0]
        y = row.iloc[1]
        dx, dy = st.session_state['deslocamentos_linhas'][i]
        dx = max(-max_deslocamento, min(max_deslocamento, dx * (x / abs(x) if x != 0 else 1)))
        dy = max(-max_deslocamento, min(max_deslocamento, dy * (y / abs(y) if y != 0 else 1)))
        pontos_linhas.append({"x": x, "y": y, "dx": dx, "dy": dy, "label": st.session_state['rotulos_linhas'][i]})

    for i, row in coordenadas_colunas.iterrows():
        if st.session_state['rotulos_colunas'][i] is None:
            continue
        x = row.iloc[0]
        y = row.iloc[1]
        dx, dy = st.session_state['deslocamentos_colunas'][i]
        dx = max(-max_deslocamento, min(max_deslocamento, dx * (x / abs(x) if x != 0 else 1)))
        dy = max(-max_deslocamento, min(max_deslocamento, dy * (y / abs(y) if y != 0 else 1)))
        pontos_colunas.append({"x": x, "y": y, "dx": dx, "dy": dy, "label": st.session_state['rotulos_colunas'][i]})

    # ----------------- MAPA ESTÁTICO (MATPLOTLIB) -----------------

    st.subheader("Mapa estático")

    fig, ax = plt.subplots(figsize=(8, 6))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    texts = []
    linha_legenda_plotada = False
    coluna_legenda_plotada = False

    for p in pontos_linhas:
        label_legenda = legenda_linhas if (mostrar_legenda and not linha_legenda_plotada) else ""
        ax.scatter(p["x"], p["y"], color="blue", marker="o", s=20, label=label_legenda)
        if mostrar_legenda and not linha_legenda_plotada:
            linha_legenda_plotada = True

        txt = ax.text(
            p["x"] + p["dx"],
            p["y"] + p["dy"],
            p["label"],
            color="blue",
            fontsize=8,
            ha="center",
            va="bottom",
            fontweight="bold",
        )
        texts.append(txt)

    for p in pontos_colunas:
        label_legenda = legenda_colunas if (mostrar_legenda and not coluna_legenda_plotada) else ""
        ax.scatter(p["x"], p["y"], color="red", marker="^", s=30, label=label_legenda)
        if mostrar_legenda and not coluna_legenda_plotada:
            coluna_legenda_plotada = True

        txt = ax.text(
            p["x"] + p["dx"],
            p["y"] + p["dy"],
            p["label"],
            color="red",
            fontsize=8,
            ha="center",
            va="bottom",
            fontweight="bold",
        )
        texts.append(txt)

    adjust_text(
        texts,
        arrowprops=None,
        force_text=(0.5, 1),
        force_points=(0.5, 1),
        expand_text=(1.2, 1.5),
        expand_points=(1.2, 1.5),
        lim=100,
    )

    ax.axhline(0, color="black", linestyle="--", linewidth=0.8)
    ax.axvline(0, color="black", linestyle="--", linewidth=0.8)
    ax.grid(color="lightgray", linestyle="--", linewidth=0.5, alpha=0.7)
    ax.set_aspect("auto")

    if mostrar_legenda and (linha_legenda_plotada or coluna_legenda_plotada):
        ax.legend(loc="upper right", frameon=False, fontsize=12)

    ax.set_title("Análise de Correspondência", fontsize=14)
    st.pyplot(fig)

    # ----------------- MAPA INTERATIVO (PLOTLY, MESMO DESIGN) -----------------

    st.subheader("Mapa interativo")

    fig_int = go.Figure()

    fig_int.add_trace(go.Scatter(
        x=[p["x"] for p in pontos_linhas],
        y=[p["y"] for p in pontos_linhas],
        mode="markers",
        name=legenda_linhas,
        marker=dict(color="blue", symbol="circle", size=8),
        showlegend=mostrar_legenda and len(pontos_linhas) > 0
    ))

    fig_int.add_trace(go.Scatter(
        x=[p["x"] for p in pontos_colunas],
        y=[p["y"] for p in pontos_colunas],
        mode="markers",
        name=legenda_colunas,
        marker=dict(color="red", symbol="triangle-up", size=10),
        showlegend=mostrar_legenda and len(pontos_colunas) > 0
    ))

    annotations = []
    for p in pontos_linhas:
        annotations.append(dict(
            x=p["x"] + p["dx"],
            y=p["y"] + p["dy"],
            xref="x",
            yref="y",
            text=p["label"],
            showarrow=False,
            font=dict(color="blue", size=10),
            xanchor="center",
            yanchor="bottom",
        ))
    for p in pontos_colunas:
        annotations.append(dict(
            x=p["x"] + p["dx"],
            y=p["y"] + p["dy"],
            xref="x",
            yref="y",
            text=p["label"],
            showarrow=False,
            font=dict(color="red", size=10),
            xanchor="center",
            yanchor="bottom",
        ))

    xs = [p["x"] for p in pontos_linhas] + [p["x"] for p in pontos_colunas]
    ys = [p["y"] for p in pontos_linhas] + [p["y"] for p in pontos_colunas]

    if xs and ys:
        x_min, x_max = min(xs), max(xs)
        y_min, y_max = min(ys), max(ys)
        dx = (x_max - x_min) * 0.1 or 0.1
        dy = (y_max - y_min) * 0.1 or 0.1
        x_range = [x_min - dx, x_max + dx]
        y_range = [y_min - dy, y_max + dy]
    else:
        x_range = [-1, 1]
        y_range = [-1, 1]

    fig_int.update_layout(
        template=None,
        title=dict(
            text="Análise de Correspondência",
            font=dict(color="black", size=14),
            x=0.5
        ),
        font=dict(color="black"),
        paper_bgcolor="white",
        plot_bgcolor="white",
        annotations=annotations,
        showlegend=mostrar_legenda,
        legend=dict(
            x=0.99, y=0.99,
            xanchor="right", yanchor="top",
            bgcolor="rgba(0,0,0,0)",
            font=dict(color="black", size=12),
        ),
        margin=dict(l=40, r=20, t=60, b=40),
        xaxis=dict(range=x_range),
        yaxis=dict(range=y_range),
    )

    fig_int.update_xaxes(
        title=dict(text="", font=dict(color="black")),
        showgrid=True,
        gridcolor="lightgray",
        gridwidth=0.5,
        griddash="dash",
        zeroline=False,
        tickfont=dict(color="black"),
        linecolor="black",
        mirror=True,
    )
    fig_int.update_yaxes(
        title=dict(text="", font=dict(color="black")),
        showgrid=True,
        gridcolor="lightgray",
        gridwidth=0.5,
        griddash="dash",
        zeroline=False,
        tickfont=dict(color="black"),
        linecolor="black",
        mirror=True,
    )

    # Linhas em 0 (como no estático)
    fig_int.add_shape(
        type="line",
        x0=x_range[0], x1=x_range[1],
        y0=0, y1=0,
        line=dict(color="black", width=1, dash="dash"),
        xref="x", yref="y",
        layer="above"
    )
    fig_int.add_shape(
        type="line",
        x0=0, x1=0,
        y0=y_range[0], y1=y_range[1],
        line=dict(color="black", width=1, dash="dash"),
        xref="x", yref="y",
        layer="above"
    )

    config = {
        "editable": True,
        "edits": {"annotationPosition": True},
    }

    st.plotly_chart(fig_int, use_container_width=True, config=config)


def cruzar_variaveis_formatada(dados):
    """
    Função para cruzar variáveis, personalizar o layout e exibir resultados em porcentagem,
    mostrando os valores das variáveis nas colunas.
    """
    st.write("### Selecione as variáveis e configure o layout")

    # Seletor de variáveis
    colunas_disponiveis = dados.columns.tolist()
    variaveis_linhas = st.multiselect("Selecione a variável para LINHAS:", colunas_disponiveis)
    variaveis_colunas = st.multiselect("Selecione as variáveis para COLUNAS:", colunas_disponiveis)

    if not variaveis_linhas or not variaveis_colunas:
        st.warning("Por favor, selecione pelo menos uma variável para linhas e colunas.")
        return

    # Configuração de agrupamento
    st.write("### Escolha como agrupar as variáveis de coluna")
    grupos = []
    for var in variaveis_colunas:
        group_name = st.text_input(f"Nome do grupo para {var}", value=f"Grupo_{var}")
        grupo_existente = next((g for g in grupos if g['nome'] == group_name), None)
        if grupo_existente:
            grupo_existente['variaveis'].append(var)
        else:
            grupos.append({'nome': group_name, 'variaveis': [var]})

    # Geração do Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Tabelas Cruzadas")

        formato_cabecalho = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Calibri',
            'font_size': 10,
            'bg_color': '#DDEBF7'
        })
        formato_celulas = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Calibri',
            'font_size': 10,
            'num_format': '0.00%'  # Exibir como porcentagem
        })

        linha_atual = 0
        for linha_var in variaveis_linhas:
            for grupo in grupos:
                titulo = f"{linha_var} x {', '.join(grupo['variaveis'])}"
                worksheet.merge_range(linha_atual, 0, linha_atual, len(grupo['variaveis']), titulo, formato_cabecalho)
                linha_atual += 1

                # Cabeçalhos detalhados com valores das colunas
                worksheet.write_string(linha_atual, 0, linha_var, formato_cabecalho)
                for idx, var in enumerate(grupo['variaveis'], start=1):
                    valores_unicos = ", ".join(dados[var].unique().astype(str))  # Listar os valores únicos
                    worksheet.write_string(linha_atual, idx, f"{var}: {valores_unicos}", formato_cabecalho)
                linha_atual += 1

                # Gerar tabela cruzada com porcentagem
                tabela_cruzada = pd.crosstab(
                    dados[linha_var],
                    [dados[v] for v in grupo['variaveis']],
                    normalize='columns'
                ) * 100

                for r_idx, (idx, row) in enumerate(tabela_cruzada.iterrows(), start=linha_atual):
                    worksheet.write(r_idx, 0, idx, formato_celulas)
                    for c_idx, value in enumerate(row, start=1):
                        worksheet.write(r_idx, c_idx, value / 100, formato_celulas)

                linha_atual += len(tabela_cruzada) + 3

    # Botão de download
    st.download_button(
        label="📥 Baixar Tabelas Cruzadas",
        data=output.getvalue(),
        file_name="tabelas_cruzadas_com_valores.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@st.cache_data
def compute_crosstab(var1, var2, data):
    """Função para calcular tabela de contingência e estatísticas associadas."""
    contingency_table = pd.crosstab(data[var1], data[var2], margins=True)
    observed = contingency_table.iloc[:-1, :-1].values
    if observed.size == 0:
        return None, None, None, None

    chi2, p, dof, expected = chi2_contingency(observed)

    col_totals = contingency_table.iloc[-1, :-1].values
    row_totals = contingency_table.iloc[:-1, -1].values
    grand_total = contingency_table.iloc[-1, -1]

    residuals_adjusted = (observed - expected) / np.sqrt(
        expected * (1 - row_totals[:, None] / grand_total) * (1 - col_totals / grand_total)
    )

    percent_table = (observed / col_totals) * 100

    return contingency_table, residuals_adjusted, percent_table, observed


import streamlit as st
import numpy as np
import pandas as pd


def hierarchical_crosstab_analysis_spss(data_tuple):
    import numpy as np

    if isinstance(data_tuple, tuple):
        data, data_exibicao = data_tuple
    else:
        data = data_tuple
        data_exibicao = data_tuple.copy()

    st.markdown("<h1 style='text-align: center; color: #0055A4;'>Análise de Crosstabs</h1>", unsafe_allow_html=True)
    st.markdown("---")

    categorical_vars = data_exibicao.select_dtypes(include=['object', 'category']).columns.tolist()

    if len(categorical_vars) < 2:
        st.warning("O banco de dados não possui variáveis suficientes para cruzamentos.")
        return

    analysis_type = st.radio("Escolha o tipo de análise:", ["Cruzamento Manual", "Cruzamento Automático"], index=0, horizontal=True)
    relevant_pairs = []

    if analysis_type == "Cruzamento Manual":
        st.markdown("### **Seleção de Variáveis para Cruzamento**")
        row_vars = st.multiselect("**Variáveis para a linha:**", categorical_vars)
        col_vars = st.multiselect("**Variáveis para a coluna:**", [var for var in categorical_vars if var not in row_vars])

        if not row_vars or not col_vars:
            st.warning("Selecione pelo menos uma variável para linha e outra para coluna.")
            return

        relevant_pairs = [(row_var, col_var) for row_var in row_vars for col_var in col_vars]

    else:
        st.markdown("### **Seleção de Variáveis para Análise Automática**")
        selected_variables = st.multiselect("**Selecione as variáveis que deseja manter na análise:**", sorted(categorical_vars), default=sorted(categorical_vars))

        if len(selected_variables) < 2:
            st.warning("Selecione pelo menos duas variáveis para cruzamento.")
            return

        if st.button("🔍 **Buscar Cruzamentos**"):
            all_pairs = [(var1, var2) for i, var1 in enumerate(selected_variables) for var2 in selected_variables[i + 1:]]
            relevant_pairs = []

            for var1, var2 in all_pairs:
                contingency_table, residuals_adjusted, _, _ = compute_crosstab(var1, var2, data)
                if contingency_table is None:
                    continue
                if np.any((residuals_adjusted > 1.9) | (residuals_adjusted < -1.9)):
                    relevant_pairs.append((var1, var2))

            if not relevant_pairs:
                st.warning("Nenhum cruzamento relevante foi encontrado com resíduos acima de 1.9 ou abaixo de -1.9.")
                return

    for var1, var2 in relevant_pairs:
        contingency_table, residuals_adjusted, percent_table, observed = compute_crosstab(var1, var2, data)
        if contingency_table is None:
            continue

        relevant_columns = [j for j in range(len(contingency_table.columns[:-1])) if np.any((residuals_adjusted[:, j] > 1.9) | (residuals_adjusted[:, j] < -1.9))]

        col_codigos_ordenados = sorted(data[var2].dropna().unique())
        col_labels = [data_exibicao.loc[data[var2] == cod, var2].iloc[0] for cod in col_codigos_ordenados]

        row_codigos_ordenados = sorted(data[var1].dropna().unique())
        row_labels = [data_exibicao.loc[data[var1] == cod, var1].iloc[0] for cod in row_codigos_ordenados]

        html_table = f"""
        <style>
            table {{ width: 90%; margin: auto; text-align: center; border-collapse: collapse; font-family: Arial; background-color: #f8f9fa; color: #222; border-radius: 8px; overflow: hidden; }}
            th {{ background-color: #0055A4; color: white; padding: 10px; font-size: 14px; border: 1px solid #ddd; }}
            td {{ border: 1px solid #ddd; padding: 8px; font-size: 13px; }}
            tr:nth-child(even) {{ background-color: #e9ecef; }}
            tr:nth-child(odd) {{ background-color: #ffffff; }}
        </style>
        <h3 style="text-align:center; color: #0055A4;">Cruzamento: {var1} x {var2}</h3>
        <table>
        <tr><th>Variável</th><th>Métrica</th>"""

        for j in relevant_columns:
            html_table += f"<th>{col_labels[j]}</th>"
        html_table += "<th>Total</th></tr>"

        for i, row_label in enumerate(row_labels[:-1]):
            total_cont = sum(observed[i, j] for j in relevant_columns)
            total_perc = sum(percent_table[i, j] for j in relevant_columns)

            html_table += f'<tr><td rowspan="3">{row_label}</td><td>Contagem</td>'
            html_table += "".join(f"<td>{observed[i, j]}</td>" for j in relevant_columns)
            html_table += f"<td style='font-weight: bold;'>{total_cont}</td></tr>"

            html_table += "<tr><td>Residuais</td>"
            html_table += "".join(
                f"<td>{residuals_adjusted[i, j]:.2f}</td>" if abs(residuals_adjusted[i, j]) > 1.9 else "<td>-</td>"
                for j in relevant_columns
            )
            html_table += "<td>-</td></tr>"

            html_table += "<tr><td>% dentro da coluna</td>"
            html_table += "".join(f"<td>{percent_table[i, j]:.2f}%</td>" for j in relevant_columns)
            html_table += f"<td style='font-weight: bold;'>{total_perc:.2f}%</td></tr>"

        html_table += "</table>"
        st.markdown(html_table, unsafe_allow_html=True)

    if data_exibicao is not None:
        st.write("### Dados Originais com Labels para Referência")
        st.dataframe(data_exibicao.head())


def calcular_correlacao_somas(dados_tuple):
    """
    Função para calcular a correlação entre variáveis após somar os valores iguais.
    Preserva os valores numéricos para cálculos e usa labels apenas para exibição.
    """

    st.title("Análise de Correlação com Agregação")

    # Garantir que estamos pegando apenas os dados numéricos (caso tenha vindo uma tupla)
    if isinstance(dados_tuple, tuple):
        dados, dados_exibicao = dados_tuple
    else:
        dados = dados_tuple
        dados_exibicao = None

    # Exibir todas as colunas para seleção
    colunas = dados.columns.tolist()
    variaveis = st.multiselect("Selecione as variáveis para cálculo de correlação:", colunas, default=colunas)

    if len(variaveis) < 2:
        st.warning("Selecione pelo menos duas variáveis para calcular a correlação.")
        return

    # Filtrar as variáveis selecionadas
    colunas_selecionadas = dados[variaveis]

    # Criar um DataFrame para agregação e conversão de strings em números
    dados_agrupados = pd.DataFrame()

    for coluna in colunas_selecionadas.columns:
        if colunas_selecionadas[coluna].dtype in ['object', 'category']:
            le = LabelEncoder()
            dados_agrupados[coluna] = le.fit_transform(colunas_selecionadas[coluna].astype(str))
        else:
            dados_agrupados[coluna] = colunas_selecionadas[coluna]

    # Preencher valores ausentes com zero
    dados_agrupados = dados_agrupados.fillna(0)

    if dados_agrupados.shape[1] < 2:
        st.warning("É necessário pelo menos duas variáveis para calcular a correlação.")
        return

    # Escolha do método de correlação
    metodo = st.radio("Escolha o método de correlação:", ["Pearson", "Spearman", "Kendall"])

    # Calcular a matriz de correlação
    matriz_correlacao = dados_agrupados.corr(method=metodo.lower())

    # Opção para ocultar a parte inferior esquerda da matriz
    ocultar_inferior_esquerda = st.checkbox("Ocultar parte inferior esquerda da matriz")

    matriz_ocultada = matriz_correlacao.copy()
    if ocultar_inferior_esquerda:
        mask = np.tril(np.ones_like(matriz_correlacao, dtype=bool), k=-1)
        matriz_ocultada[mask] = np.nan  # Oculta a parte inferior, mas mantém a diagonal

    st.write("### Matriz de Correlação")
    st.dataframe(matriz_ocultada.style.background_gradient(cmap='coolwarm', axis=None).format("{:.2f}"))

    # Gerar gráfico de calor com cores personalizadas
    fig, ax = plt.subplots(figsize=(10, 8))

    cores_personalizadas = LinearSegmentedColormap.from_list(
        "CustomMap", ["red", "white", "blue"], N=256
    )

    sns.heatmap(
        matriz_ocultada,
        annot=True,
        fmt=".2f",
        cmap=cores_personalizadas,
        vmin=-1,
        vmax=1,
        center=0,
        square=True,
        cbar_kws={"shrink": 0.8},
        mask=mask if ocultar_inferior_esquerda else None
    )

    plt.title(f"Matriz de Correlação ({metodo})", fontsize=16)
    st.pyplot(fig)

    # Salvar o gráfico como imagem em memória
    buffer = io.BytesIO()
    fig.savefig(buffer, format="png")
    buffer.seek(0)

    # Botão para download da imagem
    st.download_button(
        label="Baixar Gráfico de Correlação (PNG)",
        data=buffer,
        file_name="matriz_correlacao.png",
        mime="image/png",
    )

    # Exportação da matriz de correlação para Excel
    def salvar_matriz_excel(matriz):
        """
        Função para salvar a matriz de correlação formatada no Excel.
        """
        caminho_arquivo = "matriz_correlacao.xlsx"

        with pd.ExcelWriter(caminho_arquivo, engine="openpyxl") as writer:
            matriz.to_excel(writer, sheet_name="Correlação")

            workbook = writer.book
            sheet = writer.sheets["Correlação"]

            # Ajustar a largura das colunas
            for col_num, col_name in enumerate(matriz.columns, start=2):
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = 15

            # Criar normalizador para as cores
            norm = Normalize(vmin=-1, vmax=1)
            cmap = sns.color_palette("coolwarm", as_cmap=True)

            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=len(matriz) + 1, min_col=2, max_col=len(matriz.columns) + 1)):
                for j, cell in enumerate(row):
                    valor = cell.value
                    if ocultar_inferior_esquerda and j < i:
                        cell.value = None
                    elif isinstance(valor, (int, float)):
                        cor_rgb = cmap(norm(-valor))[:3]  # invertendo as cores
                        cor_rgb = tuple(int(c * 255) for c in cor_rgb)
                        cor_hex = "{:02X}{:02X}{:02X}".format(*cor_rgb)

                        cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
                        cell.number_format = "0.00"
                        cell.font = Font(bold=True)

        return caminho_arquivo

    st.write("### Exportar Matriz de Correlação para Excel")
    caminho_excel = salvar_matriz_excel(matriz_correlacao)
    with open(caminho_excel, "rb") as f:
        st.download_button(
            label="Baixar Matriz de Correlação (Excel)",
            data=f,
            file_name="matriz_correlacao.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# Função para calcular pesos para variáveis escolhidas
def calcular_frequencia_ponderada(df, perguntas, coluna_peso):
    """
    Função que permite escolher perguntas e calcular o peso delas usando a variável de peso.
    """
    if coluna_peso not in df.columns:
        raise ValueError(f"A coluna de pesos '{coluna_peso}' não existe no DataFrame. Colunas disponíveis: {df.columns.tolist()}")

    resultados = {}

    for pergunta in perguntas:
        if pergunta not in df.columns:
            st.warning(f"⚠️ A coluna '{pergunta}' não existe no DataFrame e será ignorada.")
            continue

        tabela_ponderada = df.groupby(pergunta)[coluna_peso].sum()
        tabela_percentual = (tabela_ponderada / tabela_ponderada.sum()) * 100

        resultados[pergunta] = {"Frequências Ponderadas": tabela_ponderada, "Percentuais Ponderados": tabela_percentual}

    return resultados


# (Trecho de exemplo mantido, mas não afeta o app principal)
data = {
    'Pergunta1': ['Sim', 'Não', 'Sim', 'Sim', 'Não', 'Não', 'Sim'],
    'Pergunta2': ['Alto', 'Médio', 'Baixo', 'Médio', 'Baixo', 'Alto', 'Alto'],
    'Peso': [1.5, 0.8, 1.2, 1.3, 1.7, 2.0, 1.1]
}
df = pd.DataFrame(data)
perguntas_selecionadas = ['Pergunta1', 'Pergunta2', 'Pergunta_Inexistente']
resultados = calcular_frequencia_ponderada(df, perguntas_selecionadas, 'Peso')
for pergunta, resultado in resultados.items():
    print(f"\n📊 Resultados para: {pergunta}")
    print("Frequências Ponderadas:")
    print(resultado["Frequências Ponderadas"])
    print("\nPercentuais Ponderados:")
    print(resultado["Percentuais Ponderados"])


# Stub simples para não quebrar quando escolher "Predição"
def modelo_predicao(dados):
    st.header("Predição")
    st.info("Módulo de predição ainda não foi implementado nesta versão do app.")


# Função principal
def main():
    uploaded_file = st.file_uploader("Carregar Arquivo", type=["csv", "xlsx", "sav"])

    if uploaded_file is not None:
        dados, dados_exibicao = carregar_dados(uploaded_file)

        if dados is not None:
            analise = st.sidebar.selectbox(
                "Escolha a análise",
                (
                    "Tratamento de Dados",
                    "Gráficos",
                    "Predição",
                    "Relatório Automatizado",
                    "Mapas de Correspondência",
                    "Relatório em Excel",
                    "Residuais",
                    "Análise de Correlação",
                    "Frequência Ponderada",
                ),
            )

            if analise == "Tratamento de Dados":
                tratamento_dados(dados)
            elif analise == "Gráficos":
                gerar_graficos(dados)
            elif analise == "Predição":
                modelo_predicao(dados)
            elif analise == "Relatório Automatizado":
                gerar_relatorio(dados)
            elif analise == "Mapas de Correspondência":
                analise_correspondencia(dados)
            elif analise == "Relatório em Excel":
                cruzar_variaveis_formatada(dados)
            elif analise == "Residuais":
                hierarchical_crosstab_analysis_spss((dados, dados_exibicao))
            elif analise == "Análise de Correlação":
                calcular_correlacao_somas(dados)
            elif analise == "Frequência Ponderada":
                colunas_disponiveis = [col for col in dados.columns if col.lower() != "peso"]
                perguntas_selecionadas = st.multiselect("Selecione as perguntas:", colunas_disponiveis)

                coluna_peso = "peso"  # ajuste se o nome da coluna de peso for diferente

                if perguntas_selecionadas:
                    resultados = calcular_frequencia_ponderada(dados, perguntas_selecionadas, coluna_peso)

                    for pergunta, resultado in resultados.items():
                        st.write(f"📊 **Resultados para: {pergunta}**")
                        st.write("Frequências Ponderadas:")
                        st.write(resultado["Frequências Ponderadas"])
                        st.write("Percentuais Ponderados:")
                        st.write(resultado["Percentuais Ponderados"])
                else:
                    st.warning("❗ Selecione pelo menos uma pergunta para calcular os pesos.")
        else:
            st.error("Falha ao carregar os dados. Verifique o formato ou o conteúdo do arquivo.")
    else:
        st.info("Por favor, carregue um arquivo para começar.")


if __name__ == "__main__":
    main()
